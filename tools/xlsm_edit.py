#!/usr/bin/env python3
"""Safe, surgical editing of the committed test workbook's worksheet cells.

Why this exists
---------------
`MEWC Lambda and VBA Unit Tests.xlsm` is the upstream/dev test workbook. Its test
*cases* live only inside that binary (there is no text form yet), so an AI tool that
adds or edits test cases has to write the ``.xlsm``. The obvious tool, openpyxl's
``save()``, is unsafe here: on a full save it silently

  * deletes ``xl/drawings/drawing1.xml`` (the Prep sheet's form-control buttons),
  * drops ``xl/richData/*`` (Prep's 8 rich values) and ``xl/metadata.xml``,
  * strips every cell's ``cm=`` dynamic-array marker (the ``@`` problem), and
  * rewrites ``styles.xml`` / ``sharedStrings.xml`` wholesale (reordered indices),

so mixing openpyxl output with original parts corrupts formatting. See CONVENTIONS.md.

What this does instead
----------------------
Pure string surgery on a copy of the ORIGINAL zip. It rewrites ONLY the individual
``<c>`` cells you name, in place, inside the target sheet's XML, and copies every
other part (styles, sharedStrings, Prep, drawings, richData, metadata, vbaProject.bin)
byte-for-byte. Consequences:

  * Prep's buttons + rich values, all styles, and the VBA are untouched.
  * Cells you do NOT edit keep their ``cm=`` markers, so the ``@`` fallout is limited
    to the cells you actually change -- your ``fix_test_formulas`` / ``lambda_update``
    VBA cleans those up when you next open the workbook in Excel.

It never uses openpyxl to *write*. openpyxl is used read-only (``read_grid``) as a
convenience for inspecting current contents.

API
---
    from xlsm_edit import apply_edits, read_grid, validate

    edits = [
        {"sheet": "biggest", "cell": "B7", "value": "new label"},
        {"sheet": "biggest", "cell": "D7", "formula": "biggest(F7:F9)"},
        {"sheet": "biggest", "cell": "H7", "value": 42},
    ]
    apply_edits(SRC, OUT, edits)          # writes OUT, leaves SRC alone

CLI
---
    py tools/xlsm_edit.py apply  <src.xlsm> <out.xlsm> <edits.json>
    py tools/xlsm_edit.py grid   <src.xlsm> <SheetName> [max_row] [max_col]
    py tools/xlsm_edit.py validate <src.xlsm> <out.xlsm> <edits.json>

``edits.json`` is a list of objects: {"sheet","cell", and one of "value"/"formula"}.
An optional "style" (integer xf index) overrides the inherited style.
"""
from __future__ import annotations

import json
import re
import shutil
import sys
import zipfile
from xml.sax.saxutils import escape


# --------------------------------------------------------------------------- #
# address helpers
# --------------------------------------------------------------------------- #
def col_to_num(letters: str) -> int:
    n = 0
    for ch in letters.upper():
        n = n * 26 + (ord(ch) - ord("A") + 1)
    return n


def split_addr(addr: str):
    m = re.fullmatch(r"([A-Za-z]+)(\d+)", addr.strip())
    if not m:
        raise ValueError(f"bad cell address: {addr!r}")
    return m.group(1).upper(), int(m.group(2)), col_to_num(m.group(1))


# --------------------------------------------------------------------------- #
# reading (openpyxl, read-only -- never writes)
# --------------------------------------------------------------------------- #
def read_grid(path, sheet, max_row=30, max_col=20):
    """Return a list of rows of (address, value) for quick inspection."""
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=False)
    ws = wb[sheet]
    out = []
    for row in ws.iter_rows(max_row=max_row, max_col=max_col):
        cells = [(c.coordinate, c.value) for c in row if c.value is not None]
        if cells:
            out.append(cells)
    wb.close()
    return out


# --------------------------------------------------------------------------- #
# zip / sheet-name plumbing
# --------------------------------------------------------------------------- #
def _sheet_part_map_xml(wb: str, rels: str):
    """{sheet display name -> 'xl/worksheets/sheetN.xml'} from raw part text."""
    relmap = dict(re.findall(r'<Relationship[^>]*Id="([^"]+)"[^>]*Target="([^"]+)"', rels))
    out = {}
    for nm, rid in re.findall(r'<sheet[^>]*name="([^"]+)"[^>]*r:id="([^"]+)"', wb):
        target = relmap.get(rid, "")
        if target and not target.startswith("/"):
            target = "xl/" + target if not target.startswith("xl/") else target
        out[_xml_unescape(nm)] = target
    return out


def _sheet_part_map(z: zipfile.ZipFile):
    """{sheet display name -> 'xl/worksheets/sheetN.xml'}."""
    return _sheet_part_map_xml(
        z.read("xl/workbook.xml").decode("utf-8", "replace"),
        z.read("xl/_rels/workbook.xml.rels").decode("utf-8", "replace"))


def _duplicate_sheets(names, data, infos, duplicates):
    """Append copies of existing sheets. duplicates: [(src name, new name)].

    A duplicated sheet is a byte copy of the source worksheet part, registered
    with fresh sheetId / r:id / part name in workbook.xml, its rels, and
    [Content_Types].xml. Copies the CELL content only: a source with its own
    rels (drawings, form controls - e.g. the Prep sheet) is rejected, since
    those referenced parts are not copied. Test sheets have no such rels.

    Mutates data/names/infos in place and returns the set of new part names.
    """
    WS_CT = "application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"
    wb = data["xl/workbook.xml"].decode("utf-8")
    rels = data["xl/_rels/workbook.xml.rels"].decode("utf-8")
    ct = data["[Content_Types].xml"].decode("utf-8")
    part_of = _sheet_part_map_xml(wb, rels)

    next_part = max(int(m) for m in re.findall(r"xl/worksheets/sheet(\d+)\.xml", " ".join(names))) + 1
    next_sheet_id = max(int(m) for m in re.findall(r'<sheet[^>]*sheetId="(\d+)"', wb)) + 1
    next_rid = max(int(m) for m in re.findall(r'Id="rId(\d+)"', rels)) + 1

    # VBA sheet code names must be unique across the workbook. A byte copy would
    # duplicate the source's <sheetPr codeName="..."> and Excel repairs the file
    # on open ("Removed Records"). Collect the used names so each copy gets a
    # fresh one.
    used_codes = set()
    for n in names:
        if re.match(r"xl/worksheets/sheet\d+\.xml$", n):
            cm = re.search(r'<sheetPr[^>]*\bcodeName="([^"]+)"', data[n].decode("utf-8"))
            if cm:
                used_codes.add(cm.group(1))
    code_ctr = [max((int(m.group(1)) for c in used_codes
                     for m in [re.fullmatch(r"Sheet(\d+)", c)] if m), default=0)]

    def _fresh_code():
        while True:
            code_ctr[0] += 1
            cand = f"Sheet{code_ctr[0]}"
            if cand not in used_codes:
                used_codes.add(cand)
                return cand

    new_parts = set()
    for src, new in duplicates:
        if src not in part_of:
            raise KeyError(f"duplicate source sheet not found: {src!r}")
        if new in part_of:
            raise KeyError(f"duplicate target name already exists: {new!r}")
        src_part = part_of[src]
        src_rels = f"xl/worksheets/_rels/{src_part.split('/')[-1]}.rels"
        if src_rels in data:
            raise ValueError(f"cannot duplicate {src!r}: it has its own rels "
                             f"({src_rels}); referenced parts would be lost")

        new_part = f"xl/worksheets/sheet{next_part}.xml"
        rid = f"rId{next_rid}"
        # Byte copy of the source worksheet, minus any selected/active markers so
        # the copy does not fight the original for the active-tab slot, and with a
        # fresh unique VBA code name so it does not collide with the source's.
        body = data[src_part].decode("utf-8")
        body = body.replace(' tabSelected="1"', "")
        src_code = re.search(r'<sheetPr[^>]*\bcodeName="([^"]+)"', body)
        if src_code:
            body = body.replace(f'codeName="{src_code.group(1)}"',
                                f'codeName="{_fresh_code()}"', 1)
        data[new_part] = body.encode("utf-8")
        infos[new_part] = None            # signal: synthesize a ZipInfo on write
        names.append(new_part)

        ct = ct.replace("</Types>",
                        f'<Override PartName="/{new_part}" ContentType="{WS_CT}"/></Types>')
        rels = rels.replace("</Relationships>",
                            f'<Relationship Id="{rid}" Type="http://schemas.openxmlformats.org/'
                            f'officeDocument/2006/relationships/worksheet" '
                            f'Target="worksheets/sheet{next_part}.xml"/></Relationships>')
        wb = re.sub(r"(</sheets>)",
                    f'<sheet name="{_xml_escape_attr(new)}" sheetId="{next_sheet_id}" '
                    f'r:id="{rid}"/>\\1', wb, count=1)

        part_of[new] = new_part
        new_parts.add(new_part)
        next_part += 1
        next_sheet_id += 1
        next_rid += 1

    data["xl/workbook.xml"] = wb.encode("utf-8")
    data["xl/_rels/workbook.xml.rels"] = rels.encode("utf-8")
    data["[Content_Types].xml"] = ct.encode("utf-8")
    return new_parts


def _xml_unescape(s: str) -> str:
    return (s.replace("&amp;", "&").replace("&lt;", "<").replace("&gt;", ">")
             .replace("&quot;", '"').replace("&apos;", "'"))


# --------------------------------------------------------------------------- #
# cell rendering
# --------------------------------------------------------------------------- #
def _render_cell(addr, col_letters, style, value=None, formula=None, array_ref=None):
    s_attr = f' s="{style}"' if style is not None else ""
    if formula is not None:
        f = formula[1:] if formula.startswith("=") else formula
        if array_ref:
            # Dynamic-array (spilling) formula: cm="1" points at the workbook's
            # XLDAPR metadata, t="array" ref=<spill range> declares the spill.
            # Without this a plain <f> that returns an array gets implicit-
            # intersected to its top-left scalar. A cached <v>0</v> is required
            # for the anchor; full-recalc-on-load fills the real value + spill.
            return (f'<c r="{addr}"{s_attr} cm="1">'
                    f'<f t="array" ref="{array_ref}">{escape(f)}</f><v>0</v></c>')
        # cached <v> is intentionally omitted; workbook is set to full-recalc on load
        return f'<c r="{addr}"{s_attr}><f>{escape(f)}</f></c>'
    if value is None:
        return f'<c r="{addr}"{s_attr}/>'
    if isinstance(value, bool):
        return f'<c r="{addr}"{s_attr} t="b"><v>{1 if value else 0}</v></c>'
    if isinstance(value, (int, float)):
        return f'<c r="{addr}"{s_attr}><v>{value!r}</v></c>' if isinstance(value, float) \
               else f'<c r="{addr}"{s_attr}><v>{value}</v></c>'
    # string -> inline string (does not touch the shared-strings table)
    return (f'<c r="{addr}"{s_attr} t="inlineStr"><is>'
            f'<t xml:space="preserve">{escape(str(value))}</t></is></c>')


_CELL_RE_CACHE = {}


def _find_cell(row_xml, addr):
    """Return (start, end) span of <c r="addr" .../> or <c ...>...</c>, or None."""
    pat = _CELL_RE_CACHE.get(addr)
    if pat is None:
        pat = re.compile(rf'<c r="{re.escape(addr)}"(?:\s[^>]*)?(?:/>|>.*?</c>)', re.S)
        _CELL_RE_CACHE[addr] = pat
    m = pat.search(row_xml)
    return (m.start(), m.end()) if m else None


def _cell_style(cell_xml):
    m = re.match(r'<c r="[A-Z]+\d+"(?:\s[^>]*?)?\ss="(\d+)"', cell_xml)
    return int(m.group(1)) if m else None


def _insert_cell_in_row(row_xml, addr, col_num, new_cell):
    """Insert new_cell into a row's cell list, keeping column order."""
    body_m = re.match(r'(<row\b[^>]*>)(.*)(</row>)', row_xml, re.S)
    if not body_m:  # self-closing empty row: <row r="n" .../>
        open_m = re.match(r'<row\b([^>]*?)/>', row_xml)
        attrs = open_m.group(1)
        return f'<row{attrs}>{new_cell}</row>'
    head, body, tail = body_m.groups()
    cells = re.findall(r'<c r="[A-Z]+\d+"(?:\s[^>]*?)?(?:/>|>.*?</c>)', body, re.S)
    out, placed = [], False
    for c in cells:
        cnum = col_to_num(re.match(r'<c r="([A-Z]+)', c).group(1))
        if not placed and col_num < cnum:
            out.append(new_cell)
            placed = True
        out.append(c)
    if not placed:
        out.append(new_cell)
    return head + "".join(out) + tail


def _row_span(sheet_xml, row_num):
    m = re.search(rf'<row r="{row_num}"(?:\s[^>]*)?(?:/>|>.*?</row>)', sheet_xml, re.S)
    return (m.start(), m.end(), m.group(0)) if m else None


def _apply_sheet_edits(sheet_xml, edits):
    """edits: list of dict(cell,value?,formula?,style?) for one sheet."""
    for e in edits:
        addr = e["cell"].upper()
        letters, rownum, colnum = split_addr(addr)
        span = _row_span(sheet_xml, rownum)
        style = e.get("style")

        # restyle: change ONLY the s= attribute, preserving cell content. An
        # absent cell becomes a styled-empty cell (so a spill target can carry
        # the column fill while still being empty enough to accept the spill).
        if "restyle" in e:
            rs = e["restyle"]
            if span is None:
                new_row = f'<row r="{rownum}"><c r="{addr}" s="{rs}"/></row>'
                sheet_xml = _insert_row(sheet_xml, rownum, new_row)
                continue
            rstart, rend, row_xml = span
            cellspan = _find_cell(row_xml, addr)
            if cellspan is None:
                new_cell = f'<c r="{addr}" s="{rs}"/>'
                new_row = _insert_cell_in_row(row_xml, addr, colnum, new_cell)
            else:
                cs, ce = cellspan
                old = row_xml[cs:ce]
                if re.match(r'<c r="[A-Z]+\d+" s="\d+"', old):
                    new_cell = re.sub(r'(<c r="[A-Z]+\d+" )s="\d+"', rf'\g<1>s="{rs}"', old, count=1)
                else:
                    new_cell = re.sub(r'(<c r="[A-Z]+\d+")', rf'\g<1> s="{rs}"', old, count=1)
                new_row = row_xml[:cs] + new_cell + row_xml[ce:]
            sheet_xml = sheet_xml[:rstart] + new_row + sheet_xml[rend:]
            continue

        if span is None:
            # need a brand-new <row>; find sheetData and insert in row order
            style_final = style
            new_cell = _render_cell(addr, letters, style_final,
                                    value=e.get("value"), formula=e.get("formula"),
                                    array_ref=e.get("array_ref"))
            new_row = f'<row r="{rownum}">{new_cell}</row>'
            sheet_xml = _insert_row(sheet_xml, rownum, new_row)
            continue

        rstart, rend, row_xml = span
        cellspan = _find_cell(row_xml, addr)
        if cellspan is not None:
            cs, ce = cellspan
            old_cell = row_xml[cs:ce]
            if style is None:
                style = _cell_style(old_cell)  # preserve existing style
            new_cell = _render_cell(addr, letters, style,
                                    value=e.get("value"), formula=e.get("formula"),
                                    array_ref=e.get("array_ref"))
            new_row = row_xml[:cs] + new_cell + row_xml[ce:]
        else:
            if style is None:
                style = _inherit_style(sheet_xml, letters, rownum)
            new_cell = _render_cell(addr, letters, style,
                                    value=e.get("value"), formula=e.get("formula"),
                                    array_ref=e.get("array_ref"))
            new_row = _insert_cell_in_row(row_xml, addr, colnum, new_cell)
        sheet_xml = sheet_xml[:rstart] + new_row + sheet_xml[rend:]
    return sheet_xml


def _inherit_style(sheet_xml, letters, rownum):
    """Best-effort style for a new cell: copy the cell directly above it."""
    for r in range(rownum - 1, 0, -1):
        span = _row_span(sheet_xml, r)
        if not span:
            continue
        cs = _find_cell(span[2], f"{letters}{r}")
        if cs:
            st = _cell_style(span[2][cs[0]:cs[1]])
            if st is not None:
                return st
        break
    return None


def _insert_row(sheet_xml, rownum, new_row):
    sd = re.search(r'(<sheetData\b[^>]*>)(.*)(</sheetData>)', sheet_xml, re.S)
    if not sd:
        raise ValueError("no <sheetData> in sheet")
    head, body, tail = sd.groups()
    rows = re.findall(r'<row\b[^>]*?(?:/>|>.*?</row>)', body, re.S)
    out, placed = [], False
    for row in rows:
        rn = int(re.match(r'<row r="(\d+)"', row).group(1))
        if not placed and rownum < rn:
            out.append(new_row)
            placed = True
        out.append(row)
    if not placed:
        out.append(new_row)
    new_body = head + "".join(out) + tail
    return sheet_xml[:sd.start()] + new_body + sheet_xml[sd.end():]


def _xml_escape_attr(s: str) -> str:
    return (s.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
             .replace('"', "&quot;"))


def _rename_sheets(wb_xml, renames):
    """Rewrite the name="" attribute of <sheet> elements in xl/workbook.xml.

    renames: {old display name -> new display name}

    This changes the sheet's display name ONLY. It does not rewrite formulas,
    defined names, or anything else that might refer to the sheet by its old
    name, so it is safe only for a sheet nothing else references yet - e.g. a
    freshly made copy. Excel would normally fix up such references itself.
    """
    seen = set()

    def repl(m):
        tag = m.group(0)
        nm = _xml_unescape(re.search(r'\bname="([^"]*)"', tag).group(1))
        if nm not in renames:
            return tag
        seen.add(nm)
        new = _xml_escape_attr(renames[nm])
        return re.sub(r'\bname="[^"]*"', f'name="{new}"', tag, count=1)

    out = re.sub(r"<sheet\b[^>]*/?>", repl, wb_xml)
    missing = sorted(set(renames) - seen)
    if missing:
        raise KeyError(f"sheet(s) not found for rename: {missing}")
    return out


def _reorder_sheets(wb_xml, order):
    """Reorder <sheet> elements in xl/workbook.xml; `order` is every sheet name.

    Tab order is the document order of <sheet> elements. Each element carries
    its own sheetId and r:id, so moving whole elements keeps every sheet bound
    to its part - nothing else needs rewriting.

    Safe here ONLY because this workbook has no sheet-local defined names:
    those carry a localSheetId that is an INDEX into this list, so reordering
    would silently reassign them. Checked and enforced below.
    """
    m = re.search(r"(<sheets>)(.*?)(</sheets>)", wb_xml, re.S)
    if not m:
        raise ValueError("no <sheets> element in workbook.xml")
    head, body, tail = m.groups()

    elems = re.findall(r"<sheet\b[^>]*/?>", body)
    by_name = {}
    for el in elems:
        nm = _xml_unescape(re.search(r'\bname="([^"]*)"', el).group(1))
        by_name[nm] = el

    if set(order) != set(by_name):
        missing = sorted(set(by_name) - set(order))
        extra = sorted(set(order) - set(by_name))
        raise ValueError(f"order must list every sheet exactly once; "
                         f"missing={missing} unknown={extra}")

    if re.search(r"localSheetId=", wb_xml):
        raise ValueError("workbook has sheet-local defined names (localSheetId); "
                         "reordering would reassign them by index")

    return wb_xml[:m.start()] + head + "".join(by_name[n] for n in order) + tail + wb_xml[m.end():]


CALC_CHAIN = "xl/calcChain.xml"


def _drop_calc_chain(names, data):
    """Remove the calculation chain part, if present.

    calcChain.xml indexes every formula cell in the workbook. The moment an
    edit adds a formula, or replaces a formula cell with a plain or empty one,
    the index disagrees with the sheets and Excel declares the file corrupt:

        Removed Records: Formula from /xl/calcChain.xml part

    It is a pure recalculation cache with no user data in it - Excel rebuilds
    it on open, and _force_full_recalc has already asked for a full recalc - so
    dropping it is both safe and the standard remedy. The part must also be
    unregistered from [Content_Types].xml and the workbook rels, or Excel will
    flag the dangling references instead.
    """
    if CALC_CHAIN not in data:
        return names, data

    names = [n for n in names if n != CALC_CHAIN]
    data = {k: v for k, v in data.items() if k != CALC_CHAIN}

    ct = data["[Content_Types].xml"].decode("utf-8")
    ct = re.sub(r'<Override[^>]*PartName="/xl/calcChain\.xml"[^>]*/>', "", ct)
    data["[Content_Types].xml"] = ct.encode("utf-8")

    rels_part = "xl/_rels/workbook.xml.rels"
    rels = data[rels_part].decode("utf-8")
    rels = re.sub(r'<Relationship[^>]*Target="calcChain\.xml"[^>]*/>', "", rels)
    data[rels_part] = rels.encode("utf-8")

    return names, data


def _force_full_recalc(wb_xml):
    """Ensure Excel recalculates on open (cached formula values were dropped)."""
    if "<calcPr" in wb_xml:
        def repl(m):
            tag = m.group(0)
            if "fullCalcOnLoad" in tag:
                return tag
            return tag[:-2] + ' fullCalcOnLoad="1"/>' if tag.endswith("/>") \
                else tag[:-1] + ' fullCalcOnLoad="1">'
        return re.sub(r'<calcPr\b[^>]*?/?>', repl, wb_xml, count=1)
    # insert a calcPr after </sheets>
    return wb_xml.replace("</sheets>", '</sheets><calcPr fullCalcOnLoad="1"/>', 1)


# --------------------------------------------------------------------------- #
# main entry point
# --------------------------------------------------------------------------- #
def apply_edits(src, out, edits, renames=None, sheet_order=None, duplicates=None):
    """Write `out` = `src` with the given cell edits applied via XML surgery.

    edits: iterable of dicts {sheet, cell, value?|formula?, style?}. `src` is
    never modified. Returns the set of sheet part-names that were touched.

    Operations are applied in a fixed order so names resolve predictably:
      1. duplicates: [(src name, new name)] - append copies of existing sheets.
      2. renames: {old name -> new name} - rename existing sheets in place.
      3. edits - reference sheets by their FINAL name (post-duplicate/rename).
      4. sheet_order: full tab order, in final names.

    renames only changes the display name (see _rename_sheets for the caveat).
    """
    by_sheet = {}
    for e in edits:
        by_sheet.setdefault(e["sheet"], []).append(e)

    with zipfile.ZipFile(src) as z:
        names = list(z.namelist())
        data = {n: z.read(n) for n in names}
        infos = {n: z.getinfo(n) for n in names}

    # 1. duplicate sheets (updates workbook.xml, rels, [Content_Types].xml)
    if duplicates:
        _duplicate_sheets(names, data, infos, duplicates)

    # 2. rename existing sheets
    wbxml = data["xl/workbook.xml"].decode("utf-8")
    if renames:
        wbxml = _rename_sheets(wbxml, renames)
    data["xl/workbook.xml"] = wbxml.encode("utf-8")

    # 3. resolve edits against the FINAL names, then apply
    part_of = _sheet_part_map_xml(
        wbxml, data["xl/_rels/workbook.xml.rels"].decode("utf-8"))
    unknown = [s for s in by_sheet if s not in part_of]
    if unknown:
        raise KeyError(f"unknown sheet(s): {unknown}. known: {sorted(part_of)}")
    touched_parts = {part_of[s] for s in by_sheet}
    for sheet, sheet_edits in by_sheet.items():
        part = part_of[sheet]
        xml = data[part].decode("utf-8")
        xml = _apply_sheet_edits(xml, sheet_edits)
        data[part] = xml.encode("utf-8")

    # 4. reorder tabs, force recalc (cached formula values were dropped)
    wbxml = data["xl/workbook.xml"].decode("utf-8")
    if sheet_order:
        wbxml = _reorder_sheets(wbxml, sheet_order)
    data["xl/workbook.xml"] = _force_full_recalc(wbxml).encode("utf-8")

    names, data = _drop_calc_chain(names, data)

    # rewrite zip, preserving order + compression; every non-edited part is byte-identical
    with zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as zo:
        for n in names:
            info = infos.get(n)
            if info is None:                       # synthesized (duplicated) part
                zi = zipfile.ZipInfo(n, date_time=(1980, 1, 1, 0, 0, 0))
                zi.compress_type = zipfile.ZIP_DEFLATED
            else:
                zi = zipfile.ZipInfo(n, date_time=info.date_time)
                zi.compress_type = info.compress_type
                zi.external_attr = info.external_attr
                zi.internal_attr = info.internal_attr
                zi.create_system = info.create_system
            zo.writestr(zi, data[n])
    return touched_parts


def validate(src, out, edits, duplicates=None, sheet_order=None):
    """Structural check: non-edited parts byte-identical, edits readable back.

    Matches the apply_edits contract: `edits` reference FINAL sheet names (after
    any duplicate/rename), so values are read back from `out` by those names.

    Legitimately-changed parts are excused: the edited sheet parts, the three
    structural parts (workbook.xml, [Content_Types].xml, workbook rels), and the
    dropped calcChain. When `duplicates` is given, newly-added worksheet parts
    are expected rather than flagged.
    """
    import openpyxl

    by_sheet = {}
    for e in edits:
        by_sheet.setdefault(e["sheet"], []).append(e)

    zs, zo = zipfile.ZipFile(src), zipfile.ZipFile(out)
    part_of_out = _sheet_part_map_xml(
        zo.read("xl/workbook.xml").decode("utf-8"),
        zo.read("xl/_rels/workbook.xml.rels").decode("utf-8"))
    touched = ({part_of_out[s] for s in by_sheet if s in part_of_out}
               | {"xl/workbook.xml", "[Content_Types].xml",
                  "xl/_rels/workbook.xml.rels"})

    report = {"ok": True, "parts_lost": [], "parts_added": [], "calc_chain_dropped": False,
              "unexpected_changes": [], "edits_verified": 0, "edits_failed": []}

    sn, on = set(zs.namelist()), set(zo.namelist())
    lost = sn - on
    report["calc_chain_dropped"] = CALC_CHAIN in lost
    report["parts_lost"] = sorted(lost - {CALC_CHAIN})

    added = on - sn
    # New worksheet parts are expected when duplicating; anything else added is not.
    expected_added = {n for n in added if re.match(r"xl/worksheets/sheet\d+\.xml$", n)} \
        if duplicates else set()
    report["parts_added"] = sorted(added - expected_added)
    if report["parts_lost"] or report["parts_added"]:
        report["ok"] = False

    for n in sorted(sn & on):
        if n in touched:
            continue
        if zs.read(n) != zo.read(n):
            report["unexpected_changes"].append(n)
            report["ok"] = False

    # read edited values back, by their final names
    wb = openpyxl.load_workbook(out, data_only=False)
    for e in edits:
        # restyle changes only the style; array_ref reads back as an ArrayFormula
        # object rather than the "=formula" string - neither round-trips through
        # this scalar-value check, so skip both.
        if "restyle" in e or "array_ref" in e:
            report["edits_verified"] += 1
            continue
        ws = wb[e["sheet"]]
        got = ws[e["cell"].upper()].value
        want = ("=" + e["formula"].lstrip("=")) if "formula" in e else e.get("value")
        ok = (str(got) == str(want)) if want is not None else (got is None)
        if ok:
            report["edits_verified"] += 1
        else:
            report["edits_failed"].append({"cell": f'{e["sheet"]}!{e["cell"]}',
                                           "want": want, "got": got})
            report["ok"] = False
    wb.close()
    return report


# --------------------------------------------------------------------------- #
# CLI
# --------------------------------------------------------------------------- #
def _main(argv):
    if len(argv) < 2:
        print(__doc__)
        return 2
    cmd = argv[1]
    if cmd == "grid":
        src, sheet = argv[2], argv[3]
        mr = int(argv[4]) if len(argv) > 4 else 30
        mc = int(argv[5]) if len(argv) > 5 else 20
        for row in read_grid(src, sheet, mr, mc):
            print("  ".join(f"{a}={v!r}" for a, v in row))
        return 0
    if cmd == "apply":
        src, out, ef = argv[2], argv[3], argv[4]
        edits = json.load(open(ef, encoding="utf-8"))
        touched = apply_edits(src, out, edits)
        print(f"applied {len(edits)} edit(s) to {len(touched)} sheet(s): {sorted(touched)}")
        return 0
    if cmd == "validate":
        src, out, ef = argv[2], argv[3], argv[4]
        edits = json.load(open(ef, encoding="utf-8"))
        rep = validate(src, out, edits)
        print(json.dumps(rep, indent=2, default=str))
        return 0 if rep["ok"] else 1
    print(f"unknown command: {cmd}")
    return 2


if __name__ == "__main__":
    sys.exit(_main(sys.argv))
